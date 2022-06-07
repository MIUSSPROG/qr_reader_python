import time
import traceback
from io import BytesIO
from tkinter.filedialog import askopenfile
from tkinter.tix import Tk

from openpyxl import Workbook, load_workbook
import cv2
import pandas as pd
from pyzbar.pyzbar import decode
import tkinter as tk
from tkinter import messagebox, CENTER, NW, SW
import imutils
from PIL import Image, ImageTk
import PIL
from datetime import date, datetime
from dateutil import parser
import requests
from urllib.request import urlopen
from PIL import ImageTk

root = tk.Tk()  # создание интерфейса
root.title('Регистрация участников')
root.geometry("800x600")  # размер интерфейса
cap = cv2.VideoCapture(0)  # объект для захвата изображения

avatarImage = tk.Label()
avatarImage.pack(side="bottom", anchor="se")
lbimage = tk.Label()  # создание области ввыода изображения с камеры
lbimage.pack(side="top", anchor="ne")  # размещение справа вверху (ne - north east)
lbText = tk.Label()  # создание лэйбла для отображения последнего пользователя
lbText.place(anchor=CENTER, relx=0.5, rely=0.5)  # размещаем по центру
lbTextLimitMessage = tk.Label()  # еще один лэйбл для учета лимита посещений
lbTextLimitMessage.place(anchor=CENTER, relx=0.5, rely=0.7)  # чуть ниже предыдущего лэйбла
good_outcome_color = '#0f0'
bad_outcome_color = "#f00"
canvas = tk.Canvas()
canvas.place(anchor=NW)
canvas.create_rectangle(10, 10, 320, 180, fill='', outline='', tags="indicator")
# canvas.create_rectangle(10, 10, 320, 180, fill="#0f0", outline="#0f0", tags="indicator")

file = 'test1.xlsx'  # название файла куда будет все сохраняться(должен быть в одном каталоге с исполняющим файлом)
timer_dist = 10*10**9
timer_dist_repeat = 10*60 # время задержки простоя
# timer_dist_repeat = 10 # время задержки между посещениями
end = time.perf_counter_ns()
# основная функция обработки данных с камеры(работает рекурсивно)
def capture():
    global end
    global canvas
    try: # для отлавливания ошибок
        xl = pd.ExcelFile(file)  # вытаскиваем метаинформацию об иксель файл
        sheets = xl.sheet_names  # вытаскиваем названия всех листов
        fio_sheet = xl.sheet_names[0]  # выделяем первый лист(в нем хранятся исходные данные(фио, лимиты посещений и т.д.))
        workbook = load_workbook(filename=file)  # открываем файл
        df = pd.read_excel(file, sheet_name=fio_sheet) # читаем полностью все данные с первой страницы
        fio_list = df.values.tolist() # преобразуем в списко(для удобства)
        fio_list_userids = [item[0] for item in fio_list]
        fio_map = {item[0]: [item[0], item[1], item[2], item[3], item[4]] for item in fio_list} # создаем словарь для удобного обращения к данным по userId

        success, frame = cap.read() # читаем данные с камры
        if success: # если камера что-то видит
            if len(decode(frame)) == 0: # если qr кода нет
                if abs(time.perf_counter_ns() - end) > timer_dist: # через 10 сек подчищаем экран
                    lbText.config(text="")
                    lbTextLimitMessage.config(text="")
                    # canvas.delete("indicator")
                    canvas.itemconfigure('indicator', fill='')
                    avatarImage.configure(image='')

            frame = imutils.resize(frame, width=320) # выводим данные с камеры в окошко шириной 320 пкс с автоматическим подгоном высоты
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB) # светокоррекция(нужно когда сжимаешь изображение в строке выше)
            img = PIL.ImageTk.PhotoImage(image=Image.fromarray(frame)) # считываем что видит камера
            lbimage.imgtk = img # сохраняем в область для отображения
            lbimage.configure(image=img)
            today = date.today() # текущая дата(в формате 13/04/2022(слэш это плохо))
            today = today.strftime("%Y/%m/%d").replace('/', '-') # меняем слэш на "-" для удобства
            for code in decode(frame): # расшифровываем что считала камера

                if today not in sheets: # проверяем наличие в иксель файле листа с текущей датов
                    workbook.create_sheet(today)
                    workbook[today]['A1'] = 'Номер'
                    workbook[today]['B1'] = 'ФИО'
                    workbook[today]['C1'] = 'Класс'
                    workbook[today]['D1'] = 'Лимит'
                    workbook[today]['E1'] = 'Фото'
                    workbook[today]['F1'] = 'Посещение'
                    workbook[today]['G1'] = 'Время'
                    workbook.save(file) # если такого нет, то создаем

                df = pd.read_excel(file, sheet_name=today) # читаем содержимое листа с текущей датой
                saved_visitors = df.values.tolist() # преобразовываем в спискок
                saved_visitors_map = {item[0]: [item[0], item[1], item[2], item[3], item[4], item[5], item[6]] for item in saved_visitors} # для удобного обращения к данным по userId делаем словарь
                saved_visitors_id = [int(item[0]) for item in saved_visitors] # список с userId

                user_id = int(code.data.decode('utf-8')) # расшифровываем qr код (там userId)

                if user_id not in saved_visitors_id: # если его еще нет в списке
                    if user_id not in fio_list_userids:
                        lbText.config(text=f"Пользователя с таким кодом не существует", font=("roboto", 30))
                        lbTextLimitMessage.config(text="")
                        canvas.itemconfigure('indicator', fill='#f00')
                    else:
                        if fio_map[user_id][3] == 0: # проверка если на листе с фио лимит 0, тогда ничего не делаем и оповещаем
                            lbTextLimitMessage.config(
                                text=f"У пользователя {fio_map[user_id][3]} нет доступа на {today}")
                            canvas.itemconfigure('indicator', fill='#f00')
                        else: # если лимит какой-то есть
                            visited = fio_map[user_id] + [1, datetime.now().strftime('%H:%M:%S')] # формируем данные с дополнительми полями(посещаемость - первый раз добавляем туда 1 и время)
                            save_user(visited, today) # сохраняем
                            lbText.config(text="Последний пользователь\n" + str(visited[1]), font=("roboto", 30)) # выводим последнего пользователя
                            lbTextLimitMessage.config(text="") # очищаем лэйбл где ограничение лимита
                            response = requests.get(fio_map[user_id][4])
                            img_data = response.content
                            img = Image.open(BytesIO(img_data)).resize((200, 200))
                            img = ImageTk.PhotoImage(img)

                            avatarImage.imgtk = img  # сохраняем в область для отображения
                            avatarImage.configure(image=img)
                            # data = urlopen(fio_map[user_id][4])
                            # image = ImageTk.PhotoImage(data=data.read())
                            # tk.Label(root, image=image).pack()
                            # time.sleep(0.5) # чтобы сразу весь лимит не использовался делаем задержку считывания с камеры qr кода в 2 секунды
                            canvas.itemconfigure('indicator', fill='#0f0')
                    end = time.perf_counter_ns()
                else: # если есть такой уже пользователь в новом листе
                    last_saved_time = parser.parse(today + " " + saved_visitors_map[user_id][6]).timestamp()
                    cur_time = datetime.now().timestamp()
                    if abs(last_saved_time - cur_time) > timer_dist_repeat:
                        if (fio_map[user_id][3] - saved_visitors_map[user_id][5]) > 0: # проверяем, хвататет ли ему лимита на очередную регистрацию на мероприятие
                            row = saved_visitors_id.index(user_id) # если хватает находим его позицию(номер строки) на листе
                            workbook[today][f'F{row + 2}'] = saved_visitors_map[user_id][5] + 1 # и в ячейку столбца "посещение" добавляем +1
                            workbook[today][f'G{row + 2}'] = datetime.now().strftime('%H:%M:%S')
                            workbook.save(filename=file) # сохраняем все
                            # time.sleep(0.5) # очередная задержка
                            lbText.config(text="Последний пользователь\n" + str(saved_visitors_map[user_id][1]), font=("roboto", 30)) # выводим последнего пользователя
                            lbTextLimitMessage.config(text="") # очищаем лэйбл где ограничение лимита
                            canvas.itemconfigure('indicator', fill='#0f0')
                            # else:
                            #     lbText.config(text="Пользователь\n" + str(saved_visitors[-1][1] + "в ожидании"),
                            #                   font=("roboto", 30))
                        else: # если не хватило лимита, то пользователь исчерпал доступный запас
                            lbTextLimitMessage.config(
                                text=f"Пользователь {saved_visitors_map[user_id][1]} исчерпал(а) лимит посещений на {today}",
                                fg='#f00')
                            lbText.config(text="Последний пользователь\n" + str(saved_visitors[-1][1]), font=("roboto", 30))
                            canvas.itemconfigure('indicator', fill='#f00')
                    end = time.perf_counter_ns()

            root.after(1, capture) # здесь происходит самовызов функции для повторного считывания, таким образом имитируя непрерывную работу камеры
    except Exception as exc: # если какая ошибка она будет перехвачени и выведена в лог
        print(exc)
        print(traceback.format_exc())


# функция сохранения пользователя в новом листе
def save_user(visited, today_date):
    workbook = load_workbook(filename=file)
    cur_sheet = workbook[today_date]
    cur_sheet.append(visited)
    workbook.save(filename=file)


root.after(1000, capture) # вызов функции capture
root.mainloop() # запуск всего этого
